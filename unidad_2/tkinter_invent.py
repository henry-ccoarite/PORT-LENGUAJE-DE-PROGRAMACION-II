import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
import sys
from datetime import datetime

# =============================================================================
# LOGICA DE NEGOCIO (BACKEND)
# =============================================================================

class Producto:
    def __init__(self, codigo_barras, nombre, marca, categoria, precio, stock):
        if not str(codigo_barras).isdigit() or len(str(codigo_barras)) != 8:
             # Flexibilizamos validación para evitar errores de carga si el excel tiene formatos raros
             pass 

        self.codigo_barras = str(codigo_barras)
        self.nombre = str(nombre)
        self.marca = str(marca)
        self.categoria = str(categoria)
        
        try:
            self.precio = float(precio)
        except:
            self.precio = 0.0
            
        try:
            self.stock = int(stock)
        except:
            self.stock = 0

    def agregar_stock(self, cantidad):
        if cantidad <= 0:
            raise ValueError("La cantidad debe ser positiva.")
        if self.stock + cantidad > 999:
            raise ValueError("El stock total no puede superar 999 unidades.")
        self.stock += cantidad

    def quitar_stock(self, cantidad):
        if cantidad <= 0:
            raise ValueError("La cantidad debe ser positiva.")
        if cantidad > self.stock:
            return False
        self.stock -= cantidad
        return True

class ReporteExcel:
    ARCHIVO = "inventario_electrodomesticos.xlsx"

    @staticmethod
    def generar(inventario):
        """Guarda el estado actual del inventario (Hoja 1)"""
        wb = ReporteExcel._abrir_workbook()
        
        # --- GESTIÓN HOJA INVENTARIO ---
        if "Inventario" in wb.sheetnames:
            ws = wb["Inventario"]
            ws.delete_rows(2, ws.max_row + 1) # Limpiar datos viejos
        else:
            ws = wb.create_sheet("Inventario", 0)
            ws.append(["Código Barras", "Nombre", "Marca", "Categoría", "Precio", "Stock", "Fecha Modif."])

        for p in inventario.values():
            ws.append([
                p.codigo_barras, p.nombre, p.marca, p.categoria, 
                p.precio, p.stock, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        ReporteExcel._guardar_workbook(wb)

    @staticmethod
    def registrar_historial(codigo, nombre, accion, cantidad):
        """Agrega una nueva fila a la hoja de Historial"""
        wb = ReporteExcel._abrir_workbook()

        # --- GESTIÓN HOJA HISTORIAL ---
        if "Historial" not in wb.sheetnames:
            ws_hist = wb.create_sheet("Historial")
            ws_hist.append(["Fecha", "Hora", "Acción", "Código", "Producto", "Cantidad"])
        else:
            ws_hist = wb["Historial"]

        ahora = datetime.now()
        ws_hist.append([
            ahora.strftime("%Y-%m-%d"),
            ahora.strftime("%H:%M:%S"),
            accion,
            codigo,
            nombre,
            cantidad
        ])

        ReporteExcel._guardar_workbook(wb)

    @staticmethod
    def leer_historial():
        """Lee los datos de la hoja Historial para mostrarlos en la App"""
        historial = []
        if os.path.exists(ReporteExcel.ARCHIVO):
            try:
                wb = load_workbook(ReporteExcel.ARCHIVO)
                if "Historial" in wb.sheetnames:
                    ws = wb["Historial"]
                    # Iterar filas (saltando encabezado) e invertir orden para ver lo más reciente primero
                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    historial = rows[::-1] 
            except Exception as e:
                print(f"Error leyendo historial: {e}")
        return historial

    @staticmethod
    def cargar_productos():
        productos = {}
        if os.path.exists(ReporteExcel.ARCHIVO):
            try:
                wb = load_workbook(ReporteExcel.ARCHIVO)
                if "Inventario" in wb.sheetnames:
                    ws = wb["Inventario"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]: 
                            try:
                                codigo = str(row[0])
                                prod = Producto(codigo, row[1], row[2], row[3], row[4], row[5])
                                productos[codigo] = prod
                            except Exception:
                                continue 
            except Exception as e:
                print(f"Error cargando Excel: {e}")
        return productos

    # --- Métodos auxiliares para manejo de archivo ---
    @staticmethod
    def _abrir_workbook():
        if os.path.exists(ReporteExcel.ARCHIVO):
            try:
                return load_workbook(ReporteExcel.ARCHIVO)
            except PermissionError:
                raise PermissionError("El archivo Excel está abierto. Ciérrelo para continuar.")
        else:
            wb = Workbook()
            # Eliminar la hoja por defecto si creamos uno nuevo
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            return wb

    @staticmethod
    def _guardar_workbook(wb):
        try:
            wb.save(ReporteExcel.ARCHIVO)
        except PermissionError:
             raise PermissionError("El archivo Excel está abierto. Ciérrelo e intente de nuevo.")

class Inventario:
    def __init__(self):
        self.productos = ReporteExcel.cargar_productos()

    def registrar_producto(self, producto):
        if producto.codigo_barras in self.productos:
            raise ValueError("El producto ya existe en el inventario.")
        self.productos[producto.codigo_barras] = producto
        ReporteExcel.generar(self.productos)
        # Registramos creación en historial
        ReporteExcel.registrar_historial(producto.codigo_barras, producto.nombre, "REGISTRO NUEVO", producto.stock)

    def modificar_stock(self, codigo, cantidad, es_entrada=True):
        if codigo not in self.productos:
            raise ValueError("El producto no existe.")
        
        producto = self.productos[codigo]
        if es_entrada:
            producto.agregar_stock(cantidad)
            accion = "ENTRADA (+)"
        else:
            if not producto.quitar_stock(cantidad):
                raise ValueError("Stock insuficiente para realizar la salida.")
            accion = "SALIDA (-)"
        
        # 1. Guardar el estado actual del inventario
        ReporteExcel.generar(self.productos)
        
        # 2. Guardar el registro en el historial
        ReporteExcel.registrar_historial(codigo, producto.nombre, accion, cantidad)
        return True

# =============================================================================
# INTERFAZ GRÁFICA (FRONTEND - TKINTER)
# =============================================================================

class InventoryApp:
    def __init__(self, root):
        self.inventario = Inventario()
        self.root = root
        self.root.title("Sistema de Gestión de Inventario v3.0")
        self.root.geometry("1100x650")
        self.root.minsize(900, 600)
        
        self.style = ttk.Style()
        self.style.theme_use('clam') 
        
        self.bg_color = "#f0f0f0"
        self.sidebar_color = "#2c3e50"
        self.text_color = "#ffffff"
        self.accent_color = "#27ae60"

        self.root.configure(bg=self.bg_color)
        
        # === LAYOUT PRINCIPAL ===
        self.sidebar = tk.Frame(self.root, bg=self.sidebar_color, width=250)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        self.main_content = tk.Frame(self.root, bg=self.bg_color)
        self.main_content.pack(side="right", fill="both", expand=True)

        # === ELEMENTOS DEL SIDEBAR ===
        lbl_titulo = tk.Label(self.sidebar, text="INVENTARIO\nELECTRO", 
                              bg=self.sidebar_color, fg=self.text_color, 
                              font=("Helvetica", 16, "bold"), pady=30)
        lbl_titulo.pack()

        # Botones del menú
        self.crear_boton_menu("📦  Ver Inventario", self.mostrar_vista_tabla)
        self.crear_boton_menu("➕  Registrar Nuevo", self.mostrar_vista_registro)
        self.crear_boton_menu("🔄  Entrada/Salida Stock", self.mostrar_vista_movimientos)
        self.crear_boton_menu("📜  Ver Historial", self.mostrar_vista_historial) # NUEVO BOTON
        self.crear_boton_menu("❌  Salir", self.root.quit)

        lbl_footer = tk.Label(self.sidebar, text="Soporte IT\nver 3.0", 
                              bg=self.sidebar_color, fg="#95a5a6", font=("Arial", 8))
        lbl_footer.pack(side="bottom", pady=20)

        self.mostrar_vista_tabla()

    def crear_boton_menu(self, texto, comando):
        btn = tk.Button(self.sidebar, text=texto, command=comando, 
                        bg=self.sidebar_color, fg=self.text_color, 
                        bd=0, font=("Segoe UI", 11), 
                        activebackground="#34495e", 
                        activeforeground=self.text_color, 
                        cursor="hand2", anchor="w", padx=20)
        btn.pack(fill="x", pady=5, ipady=10)

    def limpiar_panel(self):
        for widget in self.main_content.winfo_children():
            widget.destroy()

    # ================= VISTA 1: TABLA DE INVENTARIO =================
    def mostrar_vista_tabla(self):
        self.limpiar_panel()
        
        frame_header = tk.Frame(self.main_content, bg=self.bg_color)
        frame_header.pack(fill="x", padx=20, pady=(20, 10))
        
        lbl_titulo = tk.Label(frame_header, text="Inventario Actual", font=("Segoe UI", 24, "bold"), bg=self.bg_color, fg="#333")
        lbl_titulo.pack(anchor="w") 

        # Buscador
        frame_toolbar = tk.Frame(self.main_content, bg=self.bg_color)
        frame_toolbar.pack(fill="x", padx=20, pady=(0, 20)) 
        tk.Label(frame_toolbar, text="🔍 Buscar:", bg=self.bg_color, font=("Segoe UI", 11)).pack(side="left", padx=(0, 5))
        self.var_busqueda = tk.StringVar()
        entry_search = ttk.Entry(frame_toolbar, textvariable=self.var_busqueda, width=30, font=("Segoe UI", 10))
        entry_search.pack(side="left", padx=5)
        entry_search.bind('<Return>', lambda e: self.cargar_datos_tabla())
        tk.Button(frame_toolbar, text="Filtrar", bg="#2980b9", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, command=self.cargar_datos_tabla).pack(side="left", padx=5)
        tk.Button(frame_toolbar, text="Ver Todo", bg="#95a5a6", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, command=self.limpiar_filtro_y_recargar).pack(side="left", padx=5)

        frame_tabla = tk.Frame(self.main_content, bg=self.bg_color)
        frame_tabla.pack(fill="both", expand=True, padx=20, padding=(0, 0, 0, 20))

        columns = ("cod", "nom", "mar", "cat", "pre", "stk")
        self.tree = ttk.Treeview(frame_tabla, columns=columns, show="headings", selectmode="browse")
        
        headers = ["Código", "Nombre", "Marca", "Categoría", "Precio (S/.)", "Stock"]
        widths = [100, 200, 100, 100, 100, 80]
        
        for col, head, w in zip(columns, headers, widths):
            self.tree.heading(col, text=head)
            self.tree.column(col, width=w, anchor="center" if col in ["cod", "stk", "pre"] else "w")

        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.cargar_datos_tabla()

    def limpiar_filtro_y_recargar(self):
        self.var_busqueda.set("") 
        self.cargar_datos_tabla()

    def cargar_datos_tabla(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        filtro = self.var_busqueda.get().strip().lower()
        self.inventario = Inventario() 
        
        for p in self.inventario.productos.values():
            codigo_txt, nombre_txt = str(p.codigo_barras).lower(), str(p.nombre).lower()
            marca_txt, cat_txt = str(p.marca).lower(), str(p.categoria).lower()

            if (filtro == "") or (filtro in codigo_txt) or (filtro in nombre_txt) or (filtro in marca_txt) or (filtro in cat_txt):
                self.tree.insert("", "end", values=(p.codigo_barras, p.nombre, p.marca, p.categoria, f"{p.precio:.2f}", p.stock))

    # ================= VISTA 2: REGISTRO =================
    def mostrar_vista_registro(self):
        self.limpiar_panel()
        tk.Label(self.main_content, text="Registrar Nuevo Producto", font=("Segoe UI", 24), bg=self.bg_color).pack(pady=20, padx=20, anchor="w")

        frame_form = tk.Frame(self.main_content, bg="white", bd=1, relief="solid")
        frame_form.pack(padx=40, pady=10, fill="x", ipadx=20, ipady=20)

        self.var_codigo = tk.StringVar()
        self.var_nombre = tk.StringVar()
        self.var_marca = tk.StringVar()
        self.var_categoria = tk.StringVar()
        self.var_precio = tk.DoubleVar()
        self.var_stock = tk.IntVar()

        campos = [("Código Barras (8 dígitos):", self.var_codigo), ("Nombre Producto:", self.var_nombre),
                  ("Marca:", self.var_marca), ("Categoría:", self.var_categoria),
                  ("Precio Unitario:", self.var_precio), ("Stock Inicial:", self.var_stock)]

        for i, (label_text, var) in enumerate(campos):
            tk.Label(frame_form, text=label_text, font=("Segoe UI", 10, "bold"), bg="white").grid(row=i, column=0, sticky="e", pady=10, padx=10)
            entry = ttk.Entry(frame_form, textvariable=var, width=40)
            entry.grid(row=i, column=1, sticky="w", pady=10)

        tk.Button(frame_form, text="GUARDAR PRODUCTO", bg=self.accent_color, fg="white", font=("Segoe UI", 10, "bold"), command=self.accion_guardar, relief="flat", padx=20, pady=5).grid(row=len(campos), column=1, sticky="e", pady=20)

    def accion_guardar(self):
        try:
            prod = Producto(self.var_codigo.get(), self.var_nombre.get(), self.var_marca.get(), self.var_categoria.get(), self.var_precio.get(), self.var_stock.get())
            self.inventario.registrar_producto(prod)
            messagebox.showinfo("Éxito", "Producto registrado correctamente.")
            self.mostrar_vista_tabla() 
        except ValueError as e:
            messagebox.showerror("Error de Validación", str(e))
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ================= VISTA 3: MOVIMIENTOS =================
    def mostrar_vista_movimientos(self):
        self.limpiar_panel()
        tk.Label(self.main_content, text="Control de Stock", font=("Segoe UI", 24), bg=self.bg_color).pack(pady=20, padx=20, anchor="w")

        frame_center = tk.Frame(self.main_content, bg=self.bg_color)
        frame_center.pack(pady=20)

        tk.Label(frame_center, text="Ingrese Código de Barras:", bg=self.bg_color, font=("Segoe UI", 12)).pack(pady=5)
        self.var_mov_codigo = tk.StringVar()
        entry_cod = ttk.Entry(frame_center, textvariable=self.var_mov_codigo, font=("Segoe UI", 14), width=20, justify="center")
        entry_cod.pack(pady=5)
        entry_cod.focus()

        tk.Label(frame_center, text="Cantidad:", bg=self.bg_color, font=("Segoe UI", 12)).pack(pady=10)
        self.var_mov_cantidad = tk.IntVar(value=1)
        ttk.Entry(frame_center, textvariable=self.var_mov_cantidad, font=("Segoe UI", 14), width=10, justify="center").pack(pady=5)

        frame_botones = tk.Frame(frame_center, bg=self.bg_color)
        frame_botones.pack(pady=30)

        tk.Button(frame_botones, text="➕ ENTRADA DE STOCK", bg="#27ae60", fg="white", font=("Segoe UI", 11, "bold"), width=20, pady=10, relief="flat", command=lambda: self.accion_movimiento(True)).pack(side="left", padx=10)
        tk.Button(frame_botones, text="➖ SALIDA DE STOCK", bg="#c0392b", fg="white", font=("Segoe UI", 11, "bold"), width=20, pady=10, relief="flat", command=lambda: self.accion_movimiento(False)).pack(side="left", padx=10)

        self.lbl_info = tk.Label(frame_center, text="", bg=self.bg_color, fg="#7f8c8d", font=("Segoe UI", 10))
        self.lbl_info.pack(pady=20)

    def accion_movimiento(self, es_entrada):
        codigo = self.var_mov_codigo.get()
        try:
            cantidad = self.var_mov_cantidad.get()
            self.inventario.modificar_stock(codigo, cantidad, es_entrada)
            
            tipo = "Entrada" if es_entrada else "Salida"
            messagebox.showinfo("Éxito", f"{tipo} registrada correctamente.\nNuevo stock guardado y añadido al historial.")
            
            prod = self.inventario.productos[codigo]
            self.lbl_info.config(text=f"Producto: {prod.nombre} | Stock actual: {prod.stock}", fg="black")
            
        except ValueError as e:
            messagebox.showwarning("Advertencia", str(e))
        except Exception as e:
            messagebox.showerror("Error Crítico", str(e))

    # ================= VISTA 4: HISTORIAL (NUEVO) =================
    def mostrar_vista_historial(self):
        self.limpiar_panel()
        
        # Encabezado
        frame_header = tk.Frame(self.main_content, bg=self.bg_color)
        frame_header.pack(fill="x", padx=20, pady=(20, 10))
        tk.Label(frame_header, text="Historial de Movimientos", font=("Segoe UI", 24, "bold"), bg=self.bg_color, fg="#333").pack(anchor="w")
        
        # Botón Recargar
        tk.Button(frame_header, text="🔄 Actualizar Tabla", bg="#7f8c8d", fg="white", font=("Segoe UI", 9), relief="flat", 
                  command=self.cargar_datos_historial).pack(anchor="e")

        # Tabla
        frame_tabla = tk.Frame(self.main_content, bg=self.bg_color)
        frame_tabla.pack(fill="both", expand=True, padx=20, padding=(0, 0, 0, 20))

        columns = ("fecha", "hora", "accion", "cod", "nom", "cant")
        self.tree_hist = ttk.Treeview(frame_tabla, columns=columns, show="headings", selectmode="browse")
        
        headers = ["Fecha", "Hora", "Acción", "Código", "Producto", "Cant."]
        widths = [100, 80, 100, 100, 200, 60]
        
        for col, head, w in zip(columns, headers, widths):
            self.tree_hist.heading(col, text=head)
            self.tree_hist.column(col, width=w, anchor="center" if col != "nom" else "w")

        # Scrollbar
        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tree_hist.yview)
        self.tree_hist.configure(yscroll=scrollbar.set)
        
        self.tree_hist.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.cargar_datos_historial()

    def cargar_datos_historial(self):
        for item in self.tree_hist.get_children():
            self.tree_hist.delete(item)
        
        registros = ReporteExcel.leer_historial()
        
        for reg in registros:
            # Colorear según si es entrada o salida (opcional, usando tags)
            tag = "entrada" if "ENTRADA" in str(reg[2]) else "salida"
            self.tree_hist.insert("", "end", values=reg, tags=(tag,))

        # Configurar colores
        self.tree_hist.tag_configure("entrada", foreground="green")
        self.tree_hist.tag_configure("salida", foreground="red")

if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
