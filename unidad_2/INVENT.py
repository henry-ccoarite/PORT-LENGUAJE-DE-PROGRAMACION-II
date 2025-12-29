from openpyxl import Workbook, load_workbook
import sys
import os
from datetime import datetime


# =========================
# CLASE PRODUCTO
# =========================
class Producto:
    def __init__(
        self,
        codigo_barras: str,
        nombre: str,
        marca: str,
        categoria: str,
        precio: float,
        stock: int
    ) -> None:

        # Restricción: código de barras de 8 dígitos
        if not codigo_barras.isdigit() or len(codigo_barras) != 8:
            raise ValueError("El código de barras debe tener exactamente 8 dígitos")

        # Restricción: stock máximo 3 dígitos
        if stock < 0 or stock > 999:
            raise ValueError("El stock debe tener entre 0 y 999")

        # Restricción: precio no negativo
        if precio < 0:
            raise ValueError("El precio no puede ser negativo")

        self.codigo_barras: str = codigo_barras
        self.nombre: str = nombre
        self.marca: str = marca
        self.categoria: str = categoria
        self.precio: float = precio
        self.stock: int = stock

    def agregar_stock(self, cantidad: int) -> None:
        # Restricción: cantidad máximo 3 dígitos>A
        if cantidad <= 0 or cantidad > 999:
            raise ValueError("La cantidad debe tener entre 1 y 3 dígitos")

        if self.stock + cantidad > 999:
            raise ValueError("El stock total no puede superar los 3 dígitos")

        self.stock += cantidad

    def quitar_stock(self, cantidad: int) -> bool:
        # Restricción: cantidad máximo 3 dígitos
        if cantidad <= 0 or cantidad > 999:
            raise ValueError("La cantidad debe tener entre 1 y 3 dígitos")

        if cantidad > self.stock:
            return False

        self.stock -= cantidad
        return True


# =========================
# CLASE REPORTE EXCEL
# =========================
class ReporteExcel:
    ARCHIVO = "inventario_electrodomesticos.xlsx"

    @staticmethod
    def registrar_producto(producto) -> None:
        if os.path.exists(ReporteExcel.ARCHIVO):
            wb = load_workbook(ReporteExcel.ARCHIVO)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            ws.append([
                "Código Barras", "Nombre", "Marca",
                "Categoría", "Precio", "Stock", "Fecha"
            ])
            

        # 👉 SOLO UNA FILA
        ws.append([
            producto.codigo_barras,
            producto.nombre,
            producto.marca,
            producto.categoria,
            producto.precio,
            producto.stock,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

        wb.save(ReporteExcel.ARCHIVO)

# =========================
# CLASE INVENTARIO
# =========================
class Inventario:
    def __init__(self) -> None:
        self.productos: dict[str, Producto] = {}

    def registrar_producto(self, producto: Producto) -> None:
        # 🔒 Restricción de duplicado de códigos
        if producto.codigo_barras in self.productos:
            raise ValueError("El producto ya existe")
        
        self.productos[producto.codigo_barras] = producto
        ReporteExcel.registrar_producto(producto)

    def entrada_stock(self, codigo: str, cantidad: int) -> bool:
        if codigo in self.productos:
            self.productos[codigo].agregar_stock(cantidad)
            ReporteExcel.generar(self.productos)
            return True
        return False

    def salida_stock(self, codigo: str, cantidad: int) -> bool:
        if codigo in self.productos:
            if self.productos[codigo].quitar_stock(cantidad):
                ReporteExel.registrar_producto(producto)
                return True
        return False

    def mostrar_inventario(self) -> None:
        print("\n--- INVENTARIO DE ELECTRODOMÉSTICOS ---")
        for p in self.productos.values():
            print(f"""
Código: {p.codigo_barras}
Producto: {p.nombre}
Marca: {p.marca}
Categoría: {p.categoria}
Precio: S/. {p.precio}
Stock: {p.stock}
-----------------------------
""")


# =========================
# SISTEMA PRINCIPAL
# =========================
class SistemaInventario:
    def __init__(self) -> None:
        self.inventario = Inventario()

    def lector_codigo_barras(self) -> str:
        return input("📟 Escanee el código de barras: ")

    def menu(self) -> None:
        while True:
            print("""
🏪 SISTEMA DE INVENTARIO - TIENDA DE ELECTRODOMÉSTICOS
1. Registrar producto
2. Entrada de stock
3. Salida de stock
4. Mostrar inventario
5. Generar reporte Excel
6. Salir
""")

            opcion = input("Seleccione una opción: ")

            try:
                if opcion == "1":
                    codigo = self.lector_codigo_barras()
                    nombre = input("Nombre del producto: ")
                    marca = input("Marca: ")
                    categoria = input("Categoría (TV, Refrigeradora, Lavadora, etc): ")
                    precio = float(input("Precio: "))
                    stock = int(input("Stock inicial: "))

                    producto = Producto(
                        codigo, nombre, marca, categoria, precio, stock
                    )
                    self.inventario.registrar_producto(producto)
                    print("✅ Producto registrado correctamente")

                elif opcion == "2":
                    codigo = self.lector_codigo_barras()
                    cantidad = int(input("Cantidad a ingresar: "))
                    if self.inventario.entrada_stock(codigo, cantidad):
                        print("📦 Stock actualizado")
                    else:
                        print("❌ Producto no encontrado")

                elif opcion == "3":
                    codigo = self.lector_codigo_barras()
                    cantidad = int(input("Cantidad a retirar: "))
                    if self.inventario.salida_stock(codigo, cantidad):
                        print("🛒 Salida registrada")
                    else:
                        print("❌ Stock insuficiente o producto no encontrado")

                elif opcion == "4":
                    self.inventario.mostrar_inventario()

                elif opcion == "5":
                    #ReporteExcel.generar(self.inventario.productos)
                    print("📊 Reporte Excel actualizado")

                elif opcion == "6":
                    print("👋 Saliendo del sistema...")
                    sys.exit()

                else:
                    print("❌ Opción inválida")

            except ValueError as e:
                print(f"⚠️ Error: {e}")


# =========================
# EJECUCIÓN
# =========================
if __name__ == "__main__":
    sistema = SistemaInventario()
    sistema.menu()
